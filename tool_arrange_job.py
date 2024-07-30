# 作業者の仕事をスケジューリング
import math, datetime
import pandas as pd
import openpyxl
import win32com.client as win32

# Common
# Excelの列番（E、AAなど）をIndexにする（A=0,B=1...）
def common_get_index(_char:str):
    number = 0
    for i, char in enumerate(reversed(_char)):
        number += (ord(char) - 64) * (26 ** i)
    return number - 1

# Excel全シートの計算を実行
def calculate_excel(excel_path):
    # Excelアプリケーションを起動
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False  # Excelを表示しない

    # Excelファイルを開く
    workbook = excel.Workbooks.Open(excel_path)

    # 全シートの計算を実行
    excel.Application.CalculateFull()

    # ファイルを保存して閉じる
    workbook.Save()
    workbook.Close()

    # Excelアプリケーションを終了
    excel.Application.Quit()

# DEFULT
AVG_DAILY_STEPS = 200
Excel_path = r""
Sheet_name = ""
# Excelタイトル
SKIP_ROWS = 14
# B列　案件ID
COL_TASK = "B"
COL_TASK_N = common_get_index(COL_TASK)
# I列　Pｽﾃｯﾌﾟ数
COL_P_STEPS = "I"
COL_P_STEPS_N = common_get_index(COL_P_STEPS)
# K列　P開始予定日
COL_P_STARTDATE = "K"
COL_P_STARTDATE_N = common_get_index(COL_P_STARTDATE)
# L列　P完了予定日
COL_P_ENDDATE = "L"
COL_P_ENDDATE_N = common_get_index(COL_P_ENDDATE)
# N列　P担当者
COL_P_WORKER = "N"
COL_P_WORKER_N = common_get_index(COL_P_WORKER)
# Q列　PR担当者
COL_P_REVIEWER = "Q"
COL_P_REVIEWER_N = common_get_index(COL_P_REVIEWER)
# T列　UTｽﾃｯﾌﾟ数
COL_UT_STEPS = "T"
COL_UT_STEPS_N = common_get_index(COL_UT_STEPS)
# V列　UT開始予定日
COL_UT_STARTDATE = "V"
COL_UT_STARTDATE_N = common_get_index(COL_UT_STARTDATE)
# W列　UT完了予定日
COL_UT_ENDDATE = "W"
COL_UT_ENDDATE_N = common_get_index(COL_UT_ENDDATE)
# Y列　UT担当者
COL_UT_WORKER = "Y"
COL_UT_WORKER_N = common_get_index(COL_UT_WORKER)
# AB列　UTR担当者
COL_UT_REVIEWER = "AB"
COL_UT_REVIEWER_N = common_get_index(COL_UT_REVIEWER)

# if_weekly:今週中の作業日を数える
def get_work_days(start, end, if_weekly=False):
    count = 0
    _d = start
    while _d <= end:
        if if_weekly and _d.weekday() == 5:
            break
        if _d.weekday() < 5 and _d not in HOLIDAY_LIST:
            count += 1
        _d += datetime.timedelta(days=1)
    return count

def get_next_workday(this_day:datetime.date, if_Monday=False):
    _temp = (6 - this_day.weekday()) if if_Monday else 1
    this_day += datetime.timedelta(days=_temp)
    while(this_day.weekday()>4 or this_day in HOLIDAY_LIST):
        this_day += datetime.timedelta(days=1)
    return this_day

def get_week_date(this_day:datetime.date) -> list[datetime.date, datetime.date]:
    if this_day.weekday() > 4 or this_day in HOLIDAY_LIST:
        print("今週の日付取得失敗！")
        exit(0)

    start = end = this_day
    while start.weekday() < 5 and start not in HOLIDAY_LIST:
        start += datetime.timedelta(days=-1)
    start += datetime.timedelta(days=1)

    while end.weekday() < 5 and end not in HOLIDAY_LIST:
        end += datetime.timedelta(days=1)
    end += datetime.timedelta(days=-1)
    return start,end

def get_weekly_steps_by_worker(worker:str, this_day:datetime.date, df=None):
    if df is None:
        df = pd.read_excel(Excel_path, sheet_name=Sheet_name, skiprows=SKIP_ROWS)
    df.iloc[:, COL_P_STARTDATE_N] = pd.to_datetime(df.iloc[:, COL_P_STARTDATE_N].replace('', pd.NaT))
    df.iloc[:, COL_UT_STARTDATE_N] = pd.to_datetime(df.iloc[:, COL_UT_STARTDATE_N].replace('', pd.NaT))
    _s, _e = get_week_date(this_day)
    step = (
        # P
        df[(df.iloc[:,COL_P_WORKER_N] == worker)&(df.iloc[:,COL_P_STARTDATE_N] >= pd.to_datetime(_s))&(df.iloc[:,COL_P_STARTDATE_N] <= pd.to_datetime(_e))].iloc[:, COL_P_STEPS_N].sum()
        + df[(df.iloc[:,COL_P_REVIEWER_N] == worker)&(df.iloc[:,COL_P_STARTDATE_N] >= pd.to_datetime(_s))&(df.iloc[:,COL_P_STARTDATE_N] <= pd.to_datetime(_e))].iloc[:, COL_P_STEPS_N].sum() * 0.5
        + df[(df.iloc[:,COL_UT_WORKER_N] == worker)&(df.iloc[:,COL_UT_STARTDATE_N] >= pd.to_datetime(_s))&(df.iloc[:,COL_UT_STARTDATE_N] <= pd.to_datetime(_e))].iloc[:, COL_UT_STEPS_N].sum()
        + df[(df.iloc[:,COL_UT_REVIEWER_N] == worker)&(df.iloc[:,COL_UT_STARTDATE_N] >= pd.to_datetime(_s))&(df.iloc[:,COL_UT_STARTDATE_N] <= pd.to_datetime(_e))].iloc[:, COL_UT_STEPS_N].sum() * 0.5
    )
    print(step)
    return step
 
# 指定案件、作業者、製造かテスト（"P" or "UT"）、開始日、超過％（対平均人日）
def set_weekly_work(task, worker, p_ut, start_date, over_p, end_date=None, overwrite=False):
    calculate_excel(Excel_path)
    # データを読み込む
    df = pd.read_excel(Excel_path, sheet_name=Sheet_name, skiprows=SKIP_ROWS, engine='openpyxl')
    if p_ut == "P":
        _col_step = COL_P_STEPS_N
        _col_startdate = COL_P_STARTDATE
        _col_startdate_n = COL_P_STARTDATE_N
        _col_worker = COL_P_WORKER
    elif p_ut == "UT":
        _col_step = COL_UT_STEPS_N
        _col_startdate = COL_UT_STARTDATE
        _col_startdate_n = COL_UT_STARTDATE_N
        _col_worker = COL_UT_WORKER

    if overwrite:
        filtered_df = df[df.iloc[:, COL_TASK_N] == task]
    else:
        filtered_df = df[(df.iloc[:, COL_TASK_N] == task) & (pd.isna(df.iloc[:, _col_startdate_n]))]

    print(filtered_df)

    day_list = []
    day_counter = start_date
    while True:
        step_counter = get_weekly_steps_by_worker(worker, day_counter, df)
        avg_weekly_steps = get_work_days(day_counter, day_counter+datetime.timedelta(days=5), True) * AVG_DAILY_STEPS
        print(f'週{day_counter}のステップは{avg_weekly_steps}\t現在のステップは{step_counter}')
        over_step = over_p * avg_weekly_steps
        if step_counter < min(over_p, 1.0) * avg_weekly_steps * 0.8:
            break
        else:
            day_counter = get_next_workday(day_counter, True)
    
    for value in filtered_df.iloc[:, _col_step]:
        if value == "-":
            print("INFO: skip '-'")
            day_list.append("")
            continue
        elif math.isnan(value) :
            print(f"ERROR: 識別できないステップ{value}")
            exit(0)

        if value >= over_step:
            print(f"ERROR: over_step({over_step})過小エラーが発生！停止！(ステップ：{value})")
            exit(0)
        
        # print(value)
        step_counter += float(value)
        if step_counter >= over_step:
            day_counter = get_next_workday(day_counter, True)
            day_list.append(day_counter)
            step_counter -= avg_weekly_steps
            avg_weekly_steps = get_work_days(day_counter, day_counter+datetime.timedelta(days=5), True) * AVG_DAILY_STEPS
            over_step = over_p * avg_weekly_steps
            print(f'週{day_counter}のステップは{avg_weekly_steps}')
        elif step_counter < avg_weekly_steps:
            day_list.append(day_counter)
        else:
            day_list.append(day_counter)
            step_counter -= avg_weekly_steps
            day_counter = get_next_workday(day_counter, True)
            avg_weekly_steps = get_work_days(day_counter, day_counter+datetime.timedelta(days=5), True) * AVG_DAILY_STEPS
            over_step = over_p * avg_weekly_steps
            print(f'週{day_counter}のステップは{avg_weekly_steps}')

    print(day_list)

    print("上記日付リストで書き込みますが、よろしいでしょうか？Y/ｎ")
    if end_date is not None:
        print(f"※日付{end_date}以降は書き込まない")
    _t = input()
    if _t != "Y":
        print("中断する")
        exit(0)
        
    # Excel書き込み
    wb = openpyxl.load_workbook(Excel_path)
    ws = wb[Sheet_name]
    
    for _i in filtered_df.index.to_list():
        if end_date is not None and day_list[0] != "" and day_list[0] > end_date:
            break
        _t = day_list.pop(0)
        ws[f'{_col_startdate}{int(_i+SKIP_ROWS+2)}'] = _t
        ws[f'{_col_worker}{int(_i+SKIP_ROWS+2)}'] = worker if _t != "" else ""
    wb.save(Excel_path)

def get_reviewer(review_d:dict, worker:str):
    _list = [ [review_d[_w]['step_offset'], _w] for _w in review_d]
    _list.sort()
    while len(_list) > 0:
        reviewer = _list.pop(0)
        if reviewer[1] != worker:
            return reviewer[1]
    print("レビュアーが指定できないエラー！")
    exit(0)

# レビュー # 指定案件、作業者list、製造かテスト（"P" or "UT"）、超過％list（対平均人日）
def set_review_work(task, workers, p_ut, over_ps):
    calculate_excel(Excel_path)
    # データを読み込む
    df = pd.read_excel(Excel_path, sheet_name=Sheet_name, skiprows=SKIP_ROWS)
    filtered_df = df[df.iloc[:, COL_TASK_N] == task]
    if p_ut == "P":
        _col_startdate = COL_P_STARTDATE_N
        _col_worker = COL_P_WORKER_N
        _col_steps = COL_P_STEPS_N
        _steps_offset = float(0.5)
        _col_reviewer = COL_P_REVIEWER
    elif p_ut == "UT":
        _col_startdate = COL_UT_STARTDATE_N
        _col_worker = COL_UT_WORKER_N
        _col_steps = COL_UT_STEPS_N
        _steps_offset = float(0.25)
        _col_reviewer = COL_UT_REVIEWER
    this_day = filtered_df.iloc[:, _col_startdate].min()
    print(f"作業予定日{this_day}")
    # レビュアーの仕事量
    review_d = {}
    for i, _r in enumerate(workers):
        _t = get_weekly_steps_by_worker(_r, this_day, df)
        review_d[_r] = {
            'step': _t,
            'over_per': over_ps[i],
            'step_offset': _t / over_ps[i]
            }
    print(review_d)

    reviewer_list = []
    for _i, row in filtered_df.iterrows():
        _t = row.iloc[_col_steps]
        if _t == "-":
            print("INFO: skip '-'")
            reviewer_list.append("")
            continue

        if this_day != row.iloc[_col_startdate]:
            # print("データ更新")
            this_day = row.iloc[_col_startdate]
            for i, _r in enumerate(workers):
                _t = get_weekly_steps_by_worker(_r, this_day, df)
                review_d[_r] = {
                    'step': _t,
                    'over_per': over_ps[i],
                    'step_offset': _t / over_ps[i]
                    }
            # print(review_d)

        reviewer = get_reviewer(review_d, row.iloc[_col_worker])
        reviewer_list.append(reviewer)
        review_d[reviewer]['step'] += row.iloc[_col_steps] * _steps_offset
        review_d[reviewer]['step_offset'] = review_d[reviewer]['step'] / review_d[reviewer]['over_per']
        
    print(reviewer_list)   
    print("上記レビュアーリストで書き込みますが、よろしいでしょうか？Y/ｎ")
    _t = input()
    if _t != "Y":
        print("中断する")
        exit(0)

    # Excel書き込み
    wb = openpyxl.load_workbook(Excel_path)
    ws = wb[Sheet_name]
    
    for _i in filtered_df.index.to_list():
        ws[f'{_col_reviewer}{int(_i+SKIP_ROWS+2)}'] = reviewer_list.pop(0)
    wb.save(Excel_path)

# 週スケール⇒日スケール
def set_daily_working(worker_list):
    calculate_excel(Excel_path)
    # データを読み込む
    df = pd.read_excel(Excel_path, sheet_name=Sheet_name, skiprows=SKIP_ROWS, engine='openpyxl')
    
    # Excel書き込み
    wb = openpyxl.load_workbook(Excel_path)
    ws = wb[Sheet_name]

    for worker in worker_list:
        # P工程
        filtered_df = df[df.iloc[:, COL_P_WORKER_N] == worker]
        col_start_data = filtered_df.iloc[:, COL_P_STARTDATE_N].drop_duplicates().tolist()
        for _this_day in col_start_data:
            _t_df = filtered_df[filtered_df.iloc[:, COL_P_STARTDATE_N] == _this_day]
            _days_of_this_week = get_work_days(_this_day, _this_day+datetime.timedelta(days=5), True)
            _one_day_step = _t_df.iloc[:, COL_P_STEPS_N].dropna().sum() / float(_days_of_this_week)

            i_counter = 0
            step = 0
            for _i in _t_df.index.to_list():
                if not pd.isna(_t_df.iloc[i_counter, COL_P_STEPS_N]):
                    if i_counter == 0:
                        _start = _this_day
                    else:
                        _start = _this_day + datetime.timedelta(days=math.ceil(step / _one_day_step)-1)
                    ws[f'{COL_P_STARTDATE}{int(_i+SKIP_ROWS+2)}'] = _start
                    step += _t_df.iloc[i_counter, COL_P_STEPS_N]
                    _end = _this_day + datetime.timedelta(days=math.ceil(step / _one_day_step)-1)
                    # print(f"タスク{_i}は{_start}から{_end}")
                    ws[f'{COL_P_ENDDATE}{int(_i+SKIP_ROWS+2)}'] = _end
                i_counter += 1
        # UT工程
        filtered_df = df[df.iloc[:, COL_UT_WORKER_N] == worker]
        col_start_data = filtered_df.iloc[:, COL_UT_STARTDATE_N].drop_duplicates().tolist()
        for _this_day in col_start_data:
            if pd.isna(_this_day):
                continue
            _t_df = filtered_df[filtered_df.iloc[:, COL_UT_STARTDATE_N] == _this_day]
            _days_of_this_week = get_work_days(_this_day, _this_day+datetime.timedelta(days=5), True)
            _one_day_step = _t_df.iloc[:, COL_UT_STEPS_N].dropna().sum() / float(_days_of_this_week)

            i_counter = 0
            step = 0
            for _i in _t_df.index.to_list():
                if not pd.isna(_t_df.iloc[i_counter, COL_UT_STEPS_N]):
                    if i_counter == 0:
                        _start = _this_day
                    else:
                        _start = _this_day + datetime.timedelta(days=math.ceil(step / _one_day_step)-1)
                    ws[f'{COL_UT_STARTDATE}{int(_i+SKIP_ROWS+2)}'] = _start
                    step += _t_df.iloc[i_counter, COL_UT_STEPS_N]
                    _end = _this_day + datetime.timedelta(days=math.ceil(step / _one_day_step)-1)
                    ws[f'{COL_UT_ENDDATE}{int(_i+SKIP_ROWS+2)}'] = _end
                i_counter += 1
    wb.save(Excel_path)
    print("週スケール⇒日スケール 完了")

# 周-作業者の設定
# 休日
HOLIDAY_LIST = [
    datetime.date(2024, 8, 12),
    datetime.date(2024, 9, 16),
    datetime.date(2024, 9, 23),
    datetime.date(2024, 10, 14),
]

# a,b = get_week_date(datetime.date(2024, 8, 13))
# print(f"{a}- -{b}")

# 開始日
START_DATE = datetime.date(2024, 8, 5)
# 終了日
END_DATE = datetime.date(2024, 10, 18)
# 総工数(P + UT)　※Pレビュー、UTレビューはPの工数と同じ
TOTAL_STEPS = 23579 * 2 + 43000

# 人月
WORKER = 10

# 作業日
WORK_DAYS = get_work_days(START_DATE, END_DATE)
AVG_MONTHLY_STEPS = float(TOTAL_STEPS)/WORKER
AVG_DAILY_STEPS = AVG_MONTHLY_STEPS/20
print(f"作業日数：{WORK_DAYS}\n平均人月：{AVG_MONTHLY_STEPS}\n平均人日：{AVG_DAILY_STEPS}")

# Excel path
Excel_path = r"C:\ning\dev\tool\プログラムスケジュール【P・UT】_Test.xlsx"
Sheet_name = "プログラムスケジュール【P・UT】 "

# １.部署 # 指定案件、作業者、製造かテスト（"P" or "UT"）、開始日、超過％（対平均人日）
# set_weekly_work("CSR6-03","伍","P",datetime.date(2024, 9, 9), 1.2, datetime.date(2024, 10, 5))
# set_weekly_work("CSR6-03","劉","UT",datetime.date(2024, 10, 7), 0.8, datetime.date(2024, 10, 12))

# ２.レビュー # 指定案件、作業者list、製造かテスト（"P" or "UT"）、開始日、超過％（対平均人日）
reviewers = ["劉", "周", "寧"]
over_pers = [1.2, 0.5, 1.2]
# set_review_work("CSR6-03", reviewers, "P", over_pers)
# set_review_work("CSR6-03", reviewers, "UT", over_pers)

# ３.週スケール⇒日スケール
workers = ["劉", "範", "伍", "江", "周", "寧"]
set_daily_working(workers)
