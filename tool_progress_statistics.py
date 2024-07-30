# 進捗を統計する
import math, datetime
import pandas as pd
import openpyxl
import win32com.client as win32

# Common
# Excelの列番（E、AAなど）をIndexにする（A=0,B=1...）
def common_get_index(_char:str):
    number = 0
    for i, char in enumerate(reversed(_char)):
        number += (ord(char) - 64) * (26 ** i)
    return number - 1

# Excel全シートの計算を実行
def calculate_excel(excel_path):
    # Excelアプリケーションを起動
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False  # Excelを表示しない

    # Excelファイルを開く
    workbook = excel.Workbooks.Open(excel_path)

    # 全シートの計算を実行
    excel.Application.CalculateFull()

    # ファイルを保存して閉じる
    workbook.Save()
    workbook.Close()

    # Excelアプリケーションを終了
    excel.Application.Quit()

# DEFULT
# region
AVG_DAILY_STEPS = 200
Excel_path = r""
Sheet_name = ""
# Excelタイトル
SKIP_ROWS = 14
# B列　案件ID
COL_TASK = "B"
COL_TASK_N = common_get_index(COL_TASK)
# I列　Pｽﾃｯﾌﾟ数
COL_P_STEPS = "I"
COL_P_STEPS_N = common_get_index(COL_P_STEPS)
# K列　P開始予定日
COL_P_STARTDATE = "K"
COL_P_STARTDATE_N = common_get_index(COL_P_STARTDATE)
# L列　P完了予定日
COL_P_ENDDATE = "L"
COL_P_ENDDATE_N = common_get_index(COL_P_ENDDATE)
# P列　P完了実績日
COL_P_T_ENDDATE = "P"
COL_P_T_ENDDATE_N = common_get_index(COL_P_T_ENDDATE)
# N列　P担当者
COL_P_WORKER = "N"
COL_P_WORKER_N = common_get_index(COL_P_WORKER)
# Q列　PR担当者
COL_P_REVIEWER = "Q"
COL_P_REVIEWER_N = common_get_index(COL_P_REVIEWER)
# T列　UTｽﾃｯﾌﾟ数
COL_UT_STEPS = "T"
COL_UT_STEPS_N = common_get_index(COL_UT_STEPS)
# V列　UT開始予定日
COL_UT_STARTDATE = "V"
COL_UT_STARTDATE_N = common_get_index(COL_UT_STARTDATE)
# W列　UT完了予定日
COL_UT_ENDDATE = "W"
COL_UT_ENDDATE_N = common_get_index(COL_UT_ENDDATE)
# AA列　UT完了実績日
COL_UT_T_ENDDATE = "AA"
COL_UT_T_ENDDATE_N = common_get_index(COL_UT_T_ENDDATE)
# Y列　UT担当者
COL_UT_WORKER = "Y"
COL_UT_WORKER_N = common_get_index(COL_UT_WORKER)
# AB列　UTR担当者
COL_UT_REVIEWER = "AB"
COL_UT_REVIEWER_N = common_get_index(COL_UT_REVIEWER)
# endregion

# 報告基準日のリストを取得する
def get_day_list():
    day_list = []
    _now = START_DATE
    while _now <= END_DATE:
      day_list.append(_now)
      _now += datetime.timedelta(days=7)
    return day_list

def set_progress_by_task(day_list):
    calculate_excel(Excel_path)
    # データを読み込む
    _df = pd.read_excel(Excel_path, sheet_name=Sheet_name, skiprows=SKIP_ROWS)
    # 末尾4行を削除
    df = _df.drop(_df.index[-4:])

    _counter = 2

    # 全体
    _step_list_all = []
    for _this_day in day_list:
        # p予定/p実績/ut予定/ut実績
        _list = []
        try:
            _list.append(df[pd.to_datetime(df.iloc[:, COL_P_ENDDATE_N]) <= pd.to_datetime(_this_day)].iloc[:, COL_P_STEPS_N].dropna().sum())
        except:
            _list.append(0)
        try:
            _list.append(df[pd.to_datetime(df.iloc[:, COL_P_T_ENDDATE_N]) <= pd.to_datetime(_this_day)].iloc[:, COL_P_STEPS_N].dropna().sum())
        except:
            _list.append(0)
        try:
            _list.append(df[pd.to_datetime(df.iloc[:, COL_UT_ENDDATE_N]) <= pd.to_datetime(_this_day)].iloc[:, COL_UT_STEPS_N].dropna().sum())
        except:
            _list.append(0)
        try:
            _list.append(df[pd.to_datetime(df.iloc[:, COL_UT_T_ENDDATE_N]) <= pd.to_datetime(_this_day)].iloc[:, COL_UT_STEPS_N].dropna().sum())
        except:
            _list.append(0)
        _step_list_all.append(_list)
    print(_step_list_all)

    # Excel書き込み
    wb = openpyxl.load_workbook(Excel_path)
    if Sheet_write_name not in wb.sheetnames:
        # シートが存在しない場合、新しいシートを作成
        wb.create_sheet(title=Sheet_write_name)
        print(f"シート '{Sheet_write_name}' を新しく作成しました。")
        ws = wb[Sheet_write_name]
        ws['A1'] = "案件"
        ws['B1'] = "製造区分"
        ws['C1'] = '週'
        ws['D1'] = '予定'
        ws['E1'] = '実績'
        ws['F1'] = '比率'
    else:
        ws = wb[Sheet_write_name]

    _t_counter = 0
    for _d in _step_list_all:
        # P
        _s = _d[0]
        if _s > 0:
            _t = _d[1]
            ws[f'A{_counter}'] = "全体"
            ws[f'B{_counter}'] = "製造"
            ws[f'C{_counter}'] = day_list[_t_counter]
            ws[f'D{_counter}'] = _s
            ws[f'E{_counter}'] = _t
            ws[f'F{_counter}'] = _t/_s
            _counter += 1
        
        # UT
        _s = _d[2]
        if _s > 0:
            _t = _d[3]
            ws[f'A{_counter}'] = "全体"
            ws[f'B{_counter}'] = "UT"
            ws[f'C{_counter}'] = day_list[_t_counter]
            ws[f'D{_counter}'] = _s
            ws[f'E{_counter}'] = _t
            ws[f'F{_counter}'] = _t/_s
            _counter += 1

        _t_counter += 1

    # 案件別
    col_task = df.iloc[:, COL_TASK_N].drop_duplicates().tolist()
    for _task in col_task:
        df_task = df[df.iloc[:, COL_TASK_N] == _task]
        _step_list_task = []

        for _this_day in day_list:
            # p予定/p実績/ut予定/ut実績
            _list = []
            try:
                _list.append(df_task[pd.to_datetime(df_task.iloc[:, COL_P_ENDDATE_N]) <= pd.to_datetime(_this_day)].iloc[:, COL_P_STEPS_N].dropna().sum())
            except:
                _list.append(0)
            try:
                _list.append(df_task[pd.to_datetime(df_task.iloc[:, COL_P_T_ENDDATE_N]) <= pd.to_datetime(_this_day)].iloc[:, COL_P_STEPS_N].dropna().sum())
            except:
                _list.append(0)
            try:
                _list.append(df_task[pd.to_datetime(df_task.iloc[:, COL_UT_ENDDATE_N]) <= pd.to_datetime(_this_day)].iloc[:, COL_UT_STEPS_N].dropna().sum())
            except:
                _list.append(0)
            try:
                _list.append(df_task[pd.to_datetime(df_task.iloc[:, COL_UT_T_ENDDATE_N]) <= pd.to_datetime(_this_day)].iloc[:, COL_UT_STEPS_N].dropna().sum())
            except:
                _list.append(0)
            _step_list_task.append(_list)
        # print(_step_list_task)
        print(len(_step_list_task))

        # Excel書き込み
        _t_counter = 0
        _buffer_list = {
            "Ps": 0,
            "Pt": 0,
            "Us": 0,
            "Ut": 0
        }
        for _d in _step_list_task:
            # P
            _s = _d[0]
            if _s > 0:
                _t = _d[1]
                if _s != _buffer_list["Ps"] or _t != _buffer_list["Pt"]:
                    ws[f'A{_counter}'] = _task
                    ws[f'B{_counter}'] = "製造"
                    ws[f'C{_counter}'] = day_list[_t_counter]
                    ws[f'D{_counter}'] = _s
                    ws[f'E{_counter}'] = _t
                    ws[f'F{_counter}'] = _t/_s
                    _counter += 1
                _buffer_list["Ps"] = _s
                _buffer_list["Pt"] = _t

            # UT
            _s = _d[2]
            if _s > 0:
                _t = _d[3]
                if _s != _buffer_list["Us"] or _t != _buffer_list["Ut"]:
                    ws[f'A{_counter}'] = _task
                    ws[f'B{_counter}'] = "UT"
                    ws[f'C{_counter}'] = day_list[_t_counter]
                    ws[f'D{_counter}'] = _s
                    ws[f'E{_counter}'] = _t
                    ws[f'F{_counter}'] = _t/_s
                    _counter += 1
                _buffer_list["Us"] = _s
                _buffer_list["Ut"] = _t

            _t_counter += 1

    wb.save(Excel_path)
    print("進捗統計 完了")


# 休日
HOLIDAY_LIST = [
    datetime.date(2024, 8, 12),
    datetime.date(2024, 9, 16),
    datetime.date(2024, 9, 23),
    datetime.date(2024, 10, 14),
]

# 計上日
START_DATE = datetime.date(2024, 8, 7)
# 終了日
END_DATE = datetime.date(2024, 10, 18)

# Excel path
Excel_path = r"C:\ning\dev\tool\プログラムスケジュール【P・UT】_Test_進捗統計.xlsx"
Sheet_name = "プログラムスケジュール【P・UT】 "
Sheet_write_name = "案件別進捗"

day_list = get_day_list()
set_progress_by_task(day_list)
