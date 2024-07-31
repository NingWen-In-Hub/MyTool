# Sim conmaniesの利益計算
import numpy as np
# import matplotlib.pyplot as plt # 绘制图像いらない
import win32com.client as win32
import openpyxl

# 
# Excel path
Excel_path = r"C:\ning\dev\tool\利益計算.xlsx"

# Common
# Excel全シートの計算を実行
def calculate_excel(excel_path, if_open=False):
    # Excelアプリケーションを起動
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = not if_open  # Excelを表示しない

    # Excelファイルを開く
    workbook = excel.Workbooks.Open(excel_path)

    # 全シートの計算を実行
    excel.Application.CalculateFull()

    if not if_open:
        # ファイルを保存して閉じる
        workbook.Save()
        workbook.Close()

        # Excelアプリケーションを終了
        excel.Application.Quit()

def get_coefficients(x, y):
    # 二次拟合
    coefficients = np.polyfit(x, y, 2)
    polynomial = np.poly1d(coefficients)

    """# 生成拟合曲线
    x_fit = np.linspace(min(x), max(x), 100)
    y_fit = polynomial(x_fit)"""

    # 计算R²值
    y_pred = polynomial(x)
    ss_res = np.sum((y - y_pred) ** 2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    r_squared = 1 - (ss_res / ss_tot)

    """# 绘制图像
    plt.scatter(x, y, color='red', label='Data Points')
    plt.plot(x_fit, y_fit, color='blue', label='Quadratic Fit')
    plt.xlabel('x')
    plt.ylabel('y')
    plt.legend()
    plt.title('Quadratic Fit')
    plt.show()"""

    # 打印拟合系数和R²值
    print("Coefficients:", coefficients)
    print("R-squared:", r_squared)
    if r_squared > 0.95:
        return coefficients
    else:
        print("ERROR: 拟合不成功，需要检查数据")
        exit(0)

calculate_excel(Excel_path)

wb = openpyxl.load_workbook(Excel_path, data_only=True)
wbw = openpyxl.load_workbook(Excel_path)
sheet_list = wb.sheetnames

for sheet in sheet_list:
    # データ読み出し
    sh = wb[sheet]
    shw = wbw[sheet]
    # 指定セルの値を読み出す
    values_list = []
    num_per_min_list = []
    profit_list = []
    col_list = [chr(x) for x in range(66, 72)]
    for _col in col_list:
        values_list.append(sh[f'{_col}6'].value)
        num_per_min_list.append(sh[f'{_col}12'].value)
        profit_list.append(sh[f'{_col}10'].value)
    
    x = np.array(values_list)
    y = np.array(num_per_min_list)
    coe = get_coefficients(x, y)
    # 售价x 对 個/miny の書き込み
    shw['I2'] = f"y={coe[0]}x2+ {coe[1]}x +{coe[2]}"
    shw['I3'] = coe[0]
    shw['J3'] = coe[1]
    shw['K3'] = coe[2]

    # y = np.array(profit_list)
    # coe = get_coefficients(x, y)
    # # 售价x 对 個/miny の書き込み
    # shw['I5'] = f"y={coe[0]}x2+ {coe[1]}x +{coe[2]}"
    # shw['I6'] = coe[0]
    # shw['J6'] = coe[1]
    # shw['K6'] = coe[2]

wbw.save(Excel_path)
print("計算完了")
# calculate_excel(Excel_path, True)
